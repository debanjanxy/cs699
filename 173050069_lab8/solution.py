#!/usr/bin/python3
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math

#query number 1
def query1(df_import,df_export,wb):
	df_export_dest = (df_export.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	countries_export = df_export_dest.index
	#print(countries[0],countries[1])
	sheet1 = wb.get_sheet_by_name('Q1_exports')
	sheet1.cell(row=1,column=1).value = 'Country'
	for i in range(5):
		c = countries_export[i]
		sheet1.cell(row=i+2,column=1).value = c
	df_import_dest = (df_import.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	sheet2 = wb.get_sheet_by_name('Q1_imports')
	sheet2.cell(row=1,column=1).value = 'Country'
	countries_import = df_import_dest.index
	for i in range(5):
		c1 = countries_import[i]
		sheet2.cell(row=i+2,column=1).value = c1
	return wb

#query number 2
def query2(df_import,df_export,wb):
	df_export_commodity = (df_export.groupby('Commodity')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	commodity_export = df_export_commodity.index
	sheet3 = wb.get_sheet_by_name('Q2_exports')
	sheet3.cell(row=1,column=1).value = 'Commodity'
	for i in range(5):
		c2 = commodity_export[i]
		sheet3.cell(row=i+2,column=1).value = c2
	df_import_commodity = (df_import.groupby('Commodity')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	commodity_import = df_import_commodity.index
	sheet4 = wb.get_sheet_by_name('Q2_imports')
	sheet4.cell(row=1,column=1).value = 'Commodity'
	for i in range(5):
		c3 = commodity_import[i]
		sheet4.cell(row=i+2,column=1).value = c3
	return wb

#query number 3
def query3(df_import,df_export,wb):
	country_exports = df_export['Country'].unique()
	total_export = (df_export.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	total_import = (df_import.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	table = pd.DataFrame(columns=['Country','Total Export','Total Import','Export/Import','Trade Deficit'])
	all_countries = total_export.index
	j = 1
	for i in range(all_countries.size):
		c = all_countries[i]
		try:
			i = int(total_import.ix[c])
		except:
			i = np.nan
		try:
			e = int(total_export.ix[c])
		except:
			e = np.nan
		try:
			r = e/i
		except:
			r = np.nan
		try:
			d = e-i
		except:
			d = np.nan
		table.loc[j] = [c,e,i,r,d]
		j = j+1
	sheet5 = wb.get_sheet_by_name('Q3')
	sheet5.cell(row=1,column=1).value = 'Country'
	sheet5.cell(row=1,column=2).value = 'Total Imports'
	sheet5.cell(row=1,column=3).value = 'Total Exports'
	sheet5.cell(row=1,column=4).value = 'Export/import ratio'
	sheet5.cell(row=1,column=5).value = 'Export-import (trade deficit)'

	for i in range(all_countries.size):
		sheet5.cell(row=i+2,column=1).value = table.ix[i+1]['Country']
		if not math.isnan(table.ix[i+1]['Total Import']):
			imp = table.ix[i+1]['Total Import']
			sheet5.cell(row=i+2,column=2).value = table.ix[i+1]['Total Import']
		else:
			imp = np.nan
			sheet5.cell(row=i+2,column=2).value = 'NaN'
		if not math.isnan(table.ix[i+1]['Total Export']):
			exp = table.ix[i+1]['Total Export']
			sheet5.cell(row=i+2,column=3).value = table.ix[i+1]['Total Export']
		else:
			exp = np.nan
			sheet5.cell(row=i+2,column=3).value = 'NaN'
		if math.isnan(imp) or math.isnan(exp) or imp==0:
			sheet5.cell(row=i+2,column=4).value = 'NaN'
		else:
			sheet5.cell(row=i+2,column=4).value = table.ix[i+1]['Export/Import']
		if math.isnan(imp) or math.isnan(exp):
			sheet5.cell(row=i+2,column=5).value = 'NaN'
		else:
			sheet5.cell(row=i+2,column=5).value = table.ix[i+1]['Trade Deficit']
	return wb

#query number 4
def query(df_export,wb):
	total_export = (df_export.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))
	total_import = (df_import.groupby('Country')['Value-INR-2011-12'].agg([pd.np.sum]).sort_values('sum',ascending=False))	
	sheet6 = wb.get_sheet_by_name('Q4')
	sheet6.cell(row=1,column=1).value = 'Country'
	table_q5 = pd.DataFrame(columns=['Country','Exports','Imports'])
	j=1
	for i in range(total_export['sum'].size):
		c = total_export.index[i]
		try:
			im = int(total_import.ix[c])
		except:
			im = np.nan
		try:
			e = int(total_export.ix[c])
		except:
			e = np.nan
		if(total_export['sum'][i] >= 100000000000):
			sheet6.cell(row=j+1,column=1).value = total_export.index[i]
			table_q5.loc[j] = [c,e,im]
			j = j+1
	return wb,table_q5

#query number 5
def query5(q5_tab,wb):
	ind = q5_tab.index
	sheet7 = wb.get_sheet_by_name('Q5')
	sheet7.cell(row=1,column=1).value = 'Country'
	sheet7.cell(row=1,column=2).value = 'Import (INR)'
	sheet7.cell(row=1,column=3).value = 'Export (INR)'
	for i in range(len(ind)):
		sheet7.cell(row=i+2,column=1).value = q5_tab.ix[ind[i]]['Country']
		sheet7.cell(row=i+2,column=2).value = q5_tab.ix[ind[i]]['Imports']
		sheet7.cell(row=i+2,column=3).value = q5_tab.ix[ind[i]]['Exports']
	return wb

#query number 6
def query6(q5_tab,wb):
	q6_tab = pd.melt(q5_tab, id_vars=['Country'], value_vars=['Exports','Imports'])
	q6_table = q6_tab.sort_values(by='value',ascending=False)
	my_tab = q6_table.head(10)
	ind = my_tab.index
	sheet8 = wb.get_sheet_by_name('Q6')
	sheet8.cell(row=1,column=1).value = 'Country'
	sheet8.cell(row=1,column=2).value = 'Transaction'
	sheet8.cell(row=1,column=3).value = 'Value (INR)'
	for i in range(len(ind)):
		sheet8.cell(row=i+2,column=1).value = my_tab.ix[ind[i]]['Country']
		sheet8.cell(row=i+2,column=2).value = my_tab.ix[ind[i]]['variable']
		sheet8.cell(row=i+2,column=3).value = my_tab.ix[ind[i]]['value']
	return wb

#query number 7
def query7(df_import,df_export,wb):
	df_import_commodity = df_import['Commodity']
	df_export_commodity = df_export['Commodity']
	q7_tab = pd.Series(list(set(df_export_commodity).intersection(set(df_import_commodity))))
	q7_table = pd.DataFrame(q7_tab,columns=['Commodity'])
	sheet9 = wb.get_sheet_by_name('Q7')
	sheet9.cell(row=1,column=1).value = 'Commodity'
	for i in range(q7_table.size):
		sheet9.cell(row=i+2,column = 1).value = q7_table.ix[q7_table.index[i]]['Commodity']
	return wb
	

filename_export = 'India_Exports_2011-12_And_2012-13.xls'
df_export = pd.read_excel(filename_export)
filename_import = 'India_Imports_2011-12_And_2012-13.xls'
df_import = pd.read_excel(filename_import)
wb = openpyxl.Workbook()
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
wb.create_sheet('Q1_imports')
wb.create_sheet('Q1_exports')
wb.create_sheet('Q2_imports')
wb.create_sheet('Q2_exports')
wb.create_sheet('Q3')
wb.create_sheet('Q4')
wb.create_sheet('Q5')
wb.create_sheet('Q6')
wb.create_sheet('Q7')
sheetNames = wb.get_sheet_names()



wb = query1(df_import,df_export,wb) #query 1
wb = query2(df_import,df_export,wb) #query 2
wb = query3(df_import,df_export,wb) #query 3
wb,q5_tab = query(df_export,wb) #query 4
wb = query5(q5_tab,wb) #query 5
wb = query6(q5_tab,wb) #query 6
wb = query7(df_import,df_export,wb) #query 7
wb.save('173050069_solution.xls')