import openpyxl
import csv
import os,sys

#taking inputs
fileNameInput = sys.argv[1]
workNo = sys.argv[2]

#create csv file of the given input
fileName = fileNameInput+'.csv'
try:
	csvFile = open(fileName)
	csvReader = csv.reader(csvFile)
	csvData = list(csvReader)

	#create final result file name
	finalName = fileNameInput+'_'+workNo+'.xlsx'

	workStr = 'work no'
	headers = csvData[0]

	#get the index of the work no
	index = 0
	for i in range(len(headers)):
		head = headers[i]
		head = head.lower()	
		if workStr in head:
			index = i
			break

	#get all the rows that matches with the given work no
	result = []
	for i in range(len(csvData)):
		row = csvData[i]
		if row[index]==workNo:
			result.append(row)

	if(result):
		#create the .xlsx file where fnal result will be stored
		wb = openpyxl.Workbook()
		wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
		#create all the required sheets
		for i in range(len(result)):
			wb.create_sheet(workNo+'_'+str(i+1))

		sheetNames = wb.get_sheet_names()

		#write each result row in the sheet along with headers
		for i in range(len(result)):
			sheet = wb.get_sheet_by_name(sheetNames[i])
			for k in range(len(headers)):
				header_val = headers[k]
				sheet.cell(row=1,column=k+1).value = header_val
			result_row = result[i]
			for j in range(len(result_row)):
				val = result_row[j]
				if j==index:
					val = val+'_'+str(i+1)
				sheet.cell(row=2,column=j+1).value = val

		#save the result
		wb.save(finalName)
	else:
		print("Work no does not exist.")
except:
	print("File does not exist.")

